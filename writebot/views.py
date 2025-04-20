from django.shortcuts import render
from . import serializers
from .models import UserAction
from rest_framework import generics
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
from auth_process.views import get_keycloak_id_from_request
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


# --- API to Create Tracked Actions ---
class CreateTrackedUserActions(generics.CreateAPIView):
    queryset = UserAction.objects.all()
    serializer_class = serializers.UserActionSerializers

    def perform_create(self, serializer):
        if self.request.user.is_authenticated:
            serializer.save(user=self.request.user)
        else:
            serializer.save()


# --- Helper Functions ---
def get_element_selector(element_data):
    """Return the best locator from provided JSON — prioritize xpath and css_selector."""
    xpath = element_data.get("xpath")
    css_selector = element_data.get("css_selector")
    value = element_data.get("value", "").strip()

    if xpath:
        return f"By.XPATH, '{xpath}'"
    if css_selector:
        return f"By.CSS_SELECTOR, '{css_selector}'"

    # Fallback logic — only if no XPath or CSS is present
    attributes = element_data.get('attributes', {})
    tag = element_data.get('tagName', '').lower()
    element_id = element_data.get('id')
    name = attributes.get('name')
    classes = attributes.get('class', '').split()
    text = element_data.get('textContent', '').strip()

    if element_id:
        return f"By.ID, '{element_id}'"
    if tag == "input" and name:
        return f"By.NAME, '{name}'"
    if tag == "button" and (value or text):
        return f"By.XPATH, f'//button[contains(text(), \"{value or text}\")]'"
    if tag == "a" and (value or text):
        return f"By.XPATH, f'//a[contains(text(), \"{value or text}\")]'"
    if value or text:
        return f"By.XPATH, f'//{tag}[contains(text(), \"{value or text}\")]'"
    if tag:
        return f"By.TAG_NAME, '{tag}'"

    return None


def get_browser_configuration():
    """Get browser configuration (could be extended to use Django settings)"""
    return {
        "browser": "firefox",
        "driver_path": None,
        "implicit_wait": 10,
        "headless": False
    }


def generate_browser_setup(config):
    """Generate browser initialization lines based on configuration"""
    lines = []

    if config["browser"].lower() == "firefox":
        lines.append("from selenium.webdriver.firefox.service import Service")
        if config["driver_path"]:
            lines.append(f'browser = webdriver.Firefox(service=Service(r"{config["driver_path"]}"))')
        else:
            lines.append('browser = webdriver.Firefox()')
    elif config["browser"].lower() == "chrome":
        lines.append("from selenium.webdriver.chrome.service import Service")
        if config["driver_path"]:
            lines.append(f'browser = webdriver.Chrome(service=Service(r"{config["driver_path"]}"))')
        else:
            lines.append('browser = webdriver.Chrome()')

    lines.append(f'browser.implicitly_wait({config["implicit_wait"]})')
    if config["headless"]:
        lines.append("options = webdriver.FirefoxOptions()" if config["browser"] == "firefox"
                     else "options = webdriver.ChromeOptions()")
        lines.append("options.add_argument('--headless')")
        lines.append("browser.options = options")

    return lines


def handle_action(action, python_code):
    """Process individual action and add corresponding Selenium code"""
    action_type = action.get("type")
    element_data = action.get("element", {})
    selector = get_element_selector(element_data)

    if not selector:
        logger.warning(f"No valid selector found for action: {action}")
        return python_code

    if action_type == "click":
        python_code.append(f"    browser.find_element({selector}).click()")

    elif action_type == "input":
        value = action.get("value", "")
        python_code.append(f"    browser.find_element({selector}).clear()")
        python_code.append(f"    browser.find_element({selector}).send_keys('{value}')")

    elif action_type == "navigate":
        url = action.get("url")
        if url:
            python_code.append(f"    browser.get('{url}')")
            python_code.append(
                "    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))")

    elif action_type == "scroll":
        python_code.append(f"    browser.execute_script('window.scrollTo({action.get('x', 0)}, {action.get('y', 0)})')")

    elif action_type == "wait":
        duration = action.get("duration", 1)
        python_code.append(f"    time.sleep({duration})")

    return python_code


def add_wait_strategies(python_code):
    """Add explicit waits where needed"""
    # Add imports at the beginning
    if not any("WebDriverWait" in line for line in python_code):
        python_code.insert(0, "from selenium.webdriver.support.ui import WebDriverWait")
        python_code.insert(1, "from selenium.webdriver.support import expected_conditions as EC")

    return python_code


def flatten_actions(actions_data):
    """Flatten nested action data structure"""
    flattened_data = []
    for entry in actions_data:
        if isinstance(entry, list):
            flattened_data.extend(entry)
        else:
            flattened_data.append(entry)
    return flattened_data


# --- Load from DB ---
def get_user_actions_from_db(keycloak_id):
    try:
        actions = UserAction.objects.filter(user__keycloak_id=keycloak_id)
        actions_data = []
        for action in actions:
            if isinstance(action.json_data, str):
                actions_data.append(json.loads(action.json_data))  # If stored as string
            else:
                actions_data.append(action.json_data)  # Already a list/dict
        return actions_data
    except Exception as e:
        logger.error(f"Failed to fetch user actions for {keycloak_id}: {e}")
        return []


# --- Convert JSON to Python Code ---
def convert_json_to_python_code(json_data):
    config = get_browser_configuration()

    python_code = [
        "from selenium import webdriver",
        "from selenium.webdriver.common.by import By",
        "from selenium.webdriver.common.keys import Keys",
        "from selenium.webdriver.support.ui import WebDriverWait",
        "from selenium.webdriver.support import expected_conditions as EC",
        "import time",
        "",
        *generate_browser_setup(config),
        "",
        "try:",
    ]

    current_url = None
    last_input_element = None
    last_input_value = ""

    for action in json_data:
        action_type = action.get("type")
        element_data = action.get("element", {})
        url = action.get("url")

        # Handle URL navigation if it changes
        if url and url != current_url:
            python_code.append(f"    browser.get('{url}')")
            python_code.append(
                "    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))")
            current_url = url

        # Special handling for input fields to avoid sending each keystroke
        if action_type in ["input", "keydown"]:
            element_id = element_data.get("id")
            if element_id:
                current_value = element_data.get("value", "")

                # If we're starting a new input field or the value changed significantly
                if not last_input_element or last_input_element != element_id or len(current_value) <= len(
                        last_input_value):
                    if last_input_element and last_input_element != element_id:
                        python_code.append(
                            f"    browser.find_element(By.ID, '{last_input_element}').send_keys(Keys.TAB)")

                    python_code.append(f"    browser.find_element(By.ID, '{element_id}').clear()")
                    python_code.append(f"    browser.find_element(By.ID, '{element_id}').send_keys('{current_value}')")
                    last_input_value = current_value
                    last_input_element = element_id
                continue

        # Handle click actions
        elif action_type == "click":
            selector = get_element_selector(element_data)
            if selector:
                python_code.append(f"    browser.find_element({selector}).click()")

        # Handle submit actions
        elif action_type == "submit":
            selector = get_element_selector(element_data)
            if selector:
                python_code.append(f"    browser.find_element({selector}).submit()")

        # Add small delay between actions to simulate real user behavior
        python_code.append("    time.sleep(1)")

    python_code.extend([
        "except Exception as e:",
        "    print(f\"An error occurred: {e}\")",
        "finally:",
        "    browser.quit()",
    ])

    return '\n'.join(python_code)


# --- Download as Excel File ---
def download_python_code_excel(request):
    keycloak_id = get_keycloak_id_from_request(request)
    if not keycloak_id:
        return HttpResponse("Unauthorized", status=401)

    actions = get_user_actions_from_db(keycloak_id)
    flattened_data = flatten_actions(actions)
    python_code = convert_json_to_python_code(flattened_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Selenium Code"

    # Add metadata
    ws.append(["Generated Selenium Test Code"])
    ws.append(["Generated at:", datetime.now().isoformat()])
    ws.append(["Total actions:", len(flattened_data)])
    ws.append([])

    # Write code with line numbers
    for idx, line in enumerate(python_code.split("\n"), start=ws.max_row + 1):
        ws.cell(row=idx, column=1).value = line

    # Adjust column width
    ws.column_dimensions[get_column_letter(1)].width = 120

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=selenium_test_case.xlsx'
    return response
